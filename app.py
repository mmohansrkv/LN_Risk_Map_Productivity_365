"""
LN Risk Map — Productivity Tracker
------------------------------------------------
A Flask web app for logging daily productivity / risk-map entries per
employee and saving everything into an Excel workbook (date-named sheets).

Two kinds of accounts:
  - ADMIN (fixed credentials below): can view all dates, edit/delete any
    entry, download the Excel workbook, and manage the employee master
    list, the process list, and user login accounts.
  - USERS (email + password, created by the admin): log in and fill the
    tracker form for any employee in the master list, picking a Process
    from the admin-maintained process list. Users can view, edit, and
    delete their own submissions for the current day (no download rights,
    and they cannot edit/delete other users' entries).
    Note: the login email identifies who is filling the form, NOT whose
    data is being entered — one login can be used to log data for many
    different employees picked from the master list.

Run it with:
    pip install flask openpyxl --break-system-packages
    python app.py

Then open http://127.0.0.1:5000
Admin login:  http://127.0.0.1:5000/admin/login
User login:   http://127.0.0.1:5000/login

Excel file created at: tracker_data/productivity_tracker.xlsx
  - "Users" sheet        -> login accounts (Email, Password, Name)
  - "Master" sheet       -> employee master list (Band, Emp_Id, Emp_Name)
  - "Process_List" sheet -> list of selectable process names
  - date-named sheets    -> one row per tracker entry submitted that day

SECURITY NOTE: Credentials are stored/checked in plain text, which is
fine for a small internal/local tool. If you ever deploy this somewhere
public, swap in hashed passwords and HTTPS, and move credentials to
environment variables (SECRET_KEY / ADMIN_USERNAME / ADMIN_PASSWORD are
already read from os.environ below, with local fallbacks).
"""

import os
import io
import sqlite3
import smtplib
from email.mime.text import MIMEText
from datetime import datetime
from functools import wraps
from zoneinfo import ZoneInfo

from flask import (
    Flask, render_template_string, request, redirect, url_for,
    flash, session, send_file, abort
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_FOLDER = "tracker_data"

# All app data (users, master list, process list, tracker entries) lives in a
# real database instead of an Excel file, so it's no longer at risk of being
# wiped by a redeploy.
#
# On Render, a web service's own disk is wiped on every deploy/restart — but
# a separate Render Postgres database is persistent. Set the DATABASE_URL
# environment variable on the web service (Render dashboard -> your web
# service -> Environment) to the "Internal Database URL" shown on your
# Postgres instance's page, and the app will use it automatically.
# If DATABASE_URL isn't set (e.g. running locally), it falls back to a
# SQLite file at DB_PATH.
DATABASE_URL = os.environ.get("DATABASE_URL")
USE_POSTGRES = bool(DATABASE_URL)
DB_PATH = os.environ.get("DB_PATH", os.path.join(DATA_FOLDER, "tracker.db"))

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    # Render's internal DATABASE_URL sometimes starts with "postgres://",
    # which older/newer driver combos can be picky about; psycopg2 accepts
    # either "postgres://" or "postgresql://" fine, so no rewrite needed.

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "Mobius365")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Mobius@123")

WORK_HOURS_PER_DAY = 8.0
LEAVE_HR = 8.0  # Hr auto-filled when a user checks "On Leave today"

# The server may run in a different timezone than the users (e.g. hosting
# providers default to UTC). All dates/times shown or logged in the app use
# this timezone instead of the server's local time, so "Login: 10:10:29"
# matches what the clock on the user's wall actually says.
APP_TIMEZONE = os.environ.get("APP_TIMEZONE", "Asia/Kolkata")


def now():
    """Current datetime in APP_TIMEZONE (IST by default), not server-local time."""
    return datetime.now(ZoneInfo(APP_TIMEZONE))

# SMTP settings for the 3:00 PM "you haven't filled 8 hrs today" reminder.
# If SMTP_SERVER isn't set, reminder emails are skipped (logged only) so the
# app still runs fine locally without mail configured.
SMTP_SERVER = os.environ.get("SMTP_SERVER")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USERNAME or "noreply@example.com")
REMINDER_HOUR = int(os.environ.get("REMINDER_HOUR", "15"))   # 3:00 PM
REMINDER_MINUTE = int(os.environ.get("REMINDER_MINUTE", "0"))

USERS_SHEET = "Users"
MASTER_SHEET = "Master"
PROCESS_SHEET = "Process_List"
RESERVED_SHEETS = {USERS_SHEET, MASTER_SHEET, PROCESS_SHEET, "Info"}

USER_COLUMNS = ["Email", "Password", "Name"]
MASTER_COLUMNS = ["Band", "Emp_Id", "Emp_Name"]
# Target_Count_Hr = target units/pieces per hour for this process, used with a
# tracker row's own Count + Hr to compute "Target hr %" = Count / (Target_Count_Hr * Hr) * 100
PROCESS_COLUMNS = ["Process", "Target_Hr", "Target_Pct", "Target_Count_Hr"]

# (internal key, label shown on the form / table header)
# NOTE: "Count" is appended at the END (not next to Hr) on purpose — date
# sheets already created before this feature existed have a fixed 10-column
# layout, and appending keeps that old data readable (see read_sheet_rows /
# get_records_for_date, which zip columns positionally).
TRACKER_FIELDS = [
    ("Date", "Date"),
    ("Band", "Band"),
    ("Emp_Id", "Emp_Id"),
    ("Emp_Name", "Emp_Name"),
    ("Process", "Process"),
    ("Description", "Description"),
    ("Other", "Other"),
    ("Hr", "Hr"),
    ("Other_Description", "Description"),
    ("Logged_By", "Logged By"),
    ("Count", "Count"),
]
TRACKER_COLUMNS = [k for k, _ in TRACKER_FIELDS]

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
CELL_FONT = Font(name="Arial")


# ---------------------------------------------------------------------------
# Excel helper — still used for the "Download" buttons only
# ---------------------------------------------------------------------------
def style_header(ws, columns, widths=None):
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    if not widths:
        widths = [16] * len(columns)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# SQLite database — this is now the single source of truth for all data
# (Users, Master, Process_List, and every date's tracker entries).
# ---------------------------------------------------------------------------
def get_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        os.makedirs(DATA_FOLDER, exist_ok=True)
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
    return conn


def q(conn, sql, params=(), fetch=None):
    """Run a query against either backend. `sql` is written with SQLite-style
    '?' placeholders and translated to Postgres '%s' automatically.
    fetch: None (no result needed), "one", or "all"."""
    if USE_POSTGRES:
        sql = sql.replace("?", "%s")
        cur = conn.cursor()
        cur.execute(sql, params)
        result = cur.fetchone() if fetch == "one" else cur.fetchall() if fetch == "all" else None
        cur.close()
        return result
    else:
        cur = conn.execute(sql, params)
        return cur.fetchone() if fetch == "one" else cur.fetchall() if fetch == "all" else None


def ensure_workbook():
    """Create the database + folder + tables if they don't exist yet.
    (Kept this function's original name since every route calls it before
    touching data — only its implementation changed, from Excel to a real
    database, SQLite locally or Postgres in production.)"""
    conn = get_db()
    pk = "SERIAL PRIMARY KEY" if USE_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"
    q(conn, f"""
        CREATE TABLE IF NOT EXISTS users (
            id {pk},
            email TEXT, password TEXT, name TEXT
        )
    """)
    q(conn, f"""
        CREATE TABLE IF NOT EXISTS master (
            id {pk},
            band TEXT, emp_id TEXT, emp_name TEXT
        )
    """)
    q(conn, f"""
        CREATE TABLE IF NOT EXISTS process_list (
            id {pk},
            process TEXT, target_hr TEXT, target_pct TEXT, target_count_hr TEXT
        )
    """)
    q(conn, f"""
        CREATE TABLE IF NOT EXISTS entries (
            id {pk},
            sheet_date TEXT,
            date TEXT, band TEXT, emp_id TEXT, emp_name TEXT,
            process TEXT, description TEXT, other TEXT, hr TEXT,
            other_description TEXT, logged_by TEXT, count TEXT
        )
    """)
    q(conn, "CREATE INDEX IF NOT EXISTS idx_entries_sheet_date ON entries(sheet_date)")
    conn.commit()
    conn.close()


# db column name for each TRACKER_COLUMNS key, in the same order
_ENTRY_DB_COLS = ["date", "band", "emp_id", "emp_name", "process", "description",
                   "other", "hr", "other_description", "logged_by", "count"]


# ---------------------------------------------------------------------------
# Users (login accounts) helpers
# ---------------------------------------------------------------------------
def get_users():
    ensure_workbook()
    conn = get_db()
    rows = q(conn, "SELECT * FROM users ORDER BY id", fetch="all")
    conn.close()
    return [{"Email": r["email"], "Password": r["password"], "Name": r["name"], "_row": r["id"]} for r in rows]


def find_user_by_email(email):
    for u in get_users():
        if (u.get("Email") or "").strip().lower() == (email or "").strip().lower():
            return u
    return None


def add_user(email, password, name):
    ensure_workbook()
    conn = get_db()
    q(conn, "INSERT INTO users (email, password, name) VALUES (?, ?, ?)", (email, password, name))
    conn.commit()
    conn.close()


def update_user(row_idx, email, password, name):
    ensure_workbook()
    conn = get_db()
    q(conn, "UPDATE users SET email=?, password=?, name=? WHERE id=?", (email, password, name, row_idx))
    conn.commit()
    conn.close()


def delete_user(row_idx):
    ensure_workbook()
    conn = get_db()
    q(conn, "DELETE FROM users WHERE id=?", (row_idx,))
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Master employee list helpers
# ---------------------------------------------------------------------------
def get_master_list():
    ensure_workbook()
    conn = get_db()
    rows = q(conn, "SELECT * FROM master ORDER BY id", fetch="all")
    conn.close()
    return [{"Band": r["band"], "Emp_Id": r["emp_id"], "Emp_Name": r["emp_name"], "_row": r["id"]} for r in rows]


def add_master(data):
    ensure_workbook()
    conn = get_db()
    q(conn,
        "INSERT INTO master (band, emp_id, emp_name) VALUES (?, ?, ?)",
        (data.get("Band", ""), data.get("Emp_Id", ""), data.get("Emp_Name", "")),
    )
    conn.commit()
    conn.close()


def update_master(row_idx, data):
    ensure_workbook()
    conn = get_db()
    q(conn,
        "UPDATE master SET band=?, emp_id=?, emp_name=? WHERE id=?",
        (data.get("Band", ""), data.get("Emp_Id", ""), data.get("Emp_Name", ""), row_idx),
    )
    conn.commit()
    conn.close()


def delete_master(row_idx):
    ensure_workbook()
    conn = get_db()
    q(conn, "DELETE FROM master WHERE id=?", (row_idx,))
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Process list helpers
# ---------------------------------------------------------------------------
def get_process_list():
    ensure_workbook()
    conn = get_db()
    rows = q(conn, "SELECT * FROM process_list ORDER BY id", fetch="all")
    conn.close()
    return [{"Process": r["process"], "Target_Hr": r["target_hr"], "Target_Pct": r["target_pct"],
             "Target_Count_Hr": r["target_count_hr"], "_row": r["id"]} for r in rows]


def add_process(name, target_hr="", target_pct="", target_count_hr=""):
    ensure_workbook()
    conn = get_db()
    q(conn,
        "INSERT INTO process_list (process, target_hr, target_pct, target_count_hr) VALUES (?, ?, ?, ?)",
        (name, target_hr, target_pct, target_count_hr),
    )
    conn.commit()
    conn.close()


def update_process(row_idx, name, target_hr="", target_pct="", target_count_hr=""):
    ensure_workbook()
    conn = get_db()
    q(conn,
        "UPDATE process_list SET process=?, target_hr=?, target_pct=?, target_count_hr=? WHERE id=?",
        (name, target_hr, target_pct, target_count_hr, row_idx),
    )
    conn.commit()
    conn.close()


def delete_process(row_idx):
    ensure_workbook()
    conn = get_db()
    q(conn, "DELETE FROM process_list WHERE id=?", (row_idx,))
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Tracker entries helpers
# ---------------------------------------------------------------------------
def save_entry(data: dict, sheet_date=None):
    ensure_workbook()
    if sheet_date is None:
        sheet_date = now().strftime("%Y-%m-%d")
    conn = get_db()
    values = [data.get(col, "") for col in TRACKER_COLUMNS]
    q(conn,
        f"INSERT INTO entries (sheet_date, {', '.join(_ENTRY_DB_COLS)}) "
        f"VALUES (?, {', '.join(['?'] * len(_ENTRY_DB_COLS))})",
        [sheet_date] + values,
    )
    conn.commit()
    conn.close()


def _entry_row_to_record(row):
    record = {col: row[db_col] for col, db_col in zip(TRACKER_COLUMNS, _ENTRY_DB_COLS)}
    record["_row"] = row["id"]
    return record


def get_records_for_date(date_str):
    ensure_workbook()
    conn = get_db()
    rows = q(conn, "SELECT * FROM entries WHERE sheet_date=? ORDER BY id", (date_str,), fetch="all")
    conn.close()
    return [_entry_row_to_record(r) for r in rows]


def get_today_records():
    return get_records_for_date(now().strftime("%Y-%m-%d"))


def list_available_dates():
    ensure_workbook()
    conn = get_db()
    rows = q(conn, "SELECT DISTINCT sheet_date FROM entries", fetch="all")
    conn.close()
    dates = [r["sheet_date"] for r in rows if r["sheet_date"]]
    dates.sort(reverse=True)
    return dates


def list_available_months():
    """Distinct 'YYYY-MM' prefixes across all logged dates, newest first.
    Used to populate the monthly-download dropdown in the admin panel."""
    months = sorted(
        {d[:7] for d in list_available_dates() if len(d) >= 7},
        reverse=True,
    )
    return months


def update_record(date_str, row_idx, data: dict):
    ensure_workbook()
    conn = get_db()
    exists = q(conn, "SELECT id FROM entries WHERE id=? AND sheet_date=?", (row_idx, date_str), fetch="one")
    if not exists:
        conn.close()
        return False
    values = [data.get(col, "") for col in TRACKER_COLUMNS]
    set_clause = ", ".join(f"{db_col}=?" for db_col in _ENTRY_DB_COLS)
    q(conn, f"UPDATE entries SET {set_clause} WHERE id=?", values + [row_idx])
    conn.commit()
    conn.close()
    return True


def delete_record(date_str, row_idx):
    ensure_workbook()
    conn = get_db()
    exists = q(conn, "SELECT id FROM entries WHERE id=? AND sheet_date=?", (row_idx, date_str), fetch="one")
    if not exists:
        conn.close()
        return False
    q(conn, "DELETE FROM entries WHERE id=?", (row_idx,))
    conn.commit()
    conn.close()
    return True


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Hours worked / pending helpers
# ---------------------------------------------------------------------------
def parse_hr_total(hr_value):
    """Sum a possibly '; '-joined Hr string (from multi-process rows) into a float."""
    if not hr_value:
        return 0.0
    total = 0.0
    for part in str(hr_value).split(";"):
        part = part.strip()
        if not part:
            continue
        try:
            total += float(part)
        except ValueError:
            pass
    return total


def total_hr_for_records(records):
    return sum(parse_hr_total(r.get("Hr")) for r in records)


def pending_hr_for_records(records):
    return max(0.0, WORK_HOURS_PER_DAY - total_hr_for_records(records))


def split_process_rows(record):
    """Split a record's ';'-joined Process/Description/Hr/Count strings into a
    list of {Process, Description, Hr, Count} row dicts, for pre-filling the
    multi-row process editor on the admin edit page."""
    procs = [p.strip() for p in str(record.get("Process") or "").split(";")]
    descs = [d.strip() for d in str(record.get("Description") or "").split(";")]
    hrs = [h.strip() for h in str(record.get("Hr") or "").split(";")]
    counts = [c.strip() for c in str(record.get("Count") or "").split(";")]
    n = max(len(procs), len(descs), len(hrs), len(counts))
    rows = []
    for i in range(n):
        p = procs[i] if i < len(procs) else ""
        d = descs[i] if i < len(descs) else ""
        h = hrs[i] if i < len(hrs) else ""
        c = counts[i] if i < len(counts) else ""
        if p or d or h or c:
            rows.append({"Process": p, "Description": d, "Hr": h, "Count": c})
    if not rows:
        rows = [{"Process": "", "Description": "", "Hr": "", "Count": ""}]
    return rows


def process_breakdown(record, target_lookup=None):
    """Split a record's ';'-joined Process/Hr/Count strings into a list of
    {process, hr, pct, count, target_hr, target_pct, target_count_hr,
    target_hr_pct} dicts, where:
      - pct is that process's share of the entry's own total hours
      - target_hr/target_pct come from the admin-maintained Process List
      - target_count_hr is the admin-set units/hr target for that process
      - target_hr_pct ("Target hr %") = Count / (Target_Count_Hr * Hr) * 100,
        i.e. actual output vs. what the hourly target would expect for the
        hours actually logged on that process. Blank if no Target_Count_Hr
        is set for the process, or no hours were logged.
    Admin-only display."""
    target_lookup = target_lookup or {}
    procs = [p.strip() for p in str(record.get("Process") or "").split(";")]
    hrs = [h.strip() for h in str(record.get("Hr") or "").split(";")]
    counts = [c.strip() for c in str(record.get("Count") or "").split(";")]
    items = []
    parsed_hrs = []
    for h in hrs:
        try:
            parsed_hrs.append(float(h)) if h else parsed_hrs.append(0.0)
        except ValueError:
            parsed_hrs.append(0.0)
    parsed_counts = []
    for c in counts:
        try:
            parsed_counts.append(float(c)) if c else parsed_counts.append(0.0)
        except ValueError:
            parsed_counts.append(0.0)
    total = sum(parsed_hrs)
    for i, proc in enumerate(procs):
        if not proc:
            continue
        hr = parsed_hrs[i] if i < len(parsed_hrs) else 0.0
        count = parsed_counts[i] if i < len(parsed_counts) else 0.0
        pct = round((hr / total) * 100) if total > 0 else 0
        target = target_lookup.get(proc, {})
        target_count_hr = target.get("Target_Count_Hr") or ""
        target_hr_pct = ""
        try:
            tch = float(target_count_hr)
            if tch > 0 and hr > 0:
                target_hr_pct = round((count / (tch * hr)) * 100)
        except (ValueError, TypeError):
            target_hr_pct = ""
        items.append({
            "process": proc, "hr": hr, "pct": pct, "count": count,
            "target_hr": target.get("Target_Hr") or "",
            "target_pct": target.get("Target_Pct") or "",
            "target_count_hr": target_count_hr,
            "target_hr_pct": target_hr_pct,
        })
    return items


def expand_record_rows(record, target_lookup=None):
    """Split a record's ';'-joined Process/Description/Hr/Count strings into
    one dict per process, each a full copy of the record with Process/
    Description/Hr/Count overridden to that single process's values (instead
    of the combined semicolon string), plus the same pct/target figures as
    process_breakdown. Used to render one table row per process instead of
    squashing multiple processes into a single semicolon-joined cell.

    Each returned dict also carries:
      - _is_first: True for the first process row of this entry (used to
        only print the shared, non-process columns and Actions once, with
        a rowspan, instead of repeating them on every process sub-row)
      - _group_size: how many process rows this entry expanded into
    """
    target_lookup = target_lookup or {}
    procs = [p.strip() for p in str(record.get("Process") or "").split(";")]
    descs = [d.strip() for d in str(record.get("Description") or "").split(";")]
    hrs = [h.strip() for h in str(record.get("Hr") or "").split(";")]
    counts = [c.strip() for c in str(record.get("Count") or "").split(";")]
    n = max(len(procs), len(descs), len(hrs), len(counts))

    raw = []
    for i in range(n):
        p = procs[i] if i < len(procs) else ""
        d = descs[i] if i < len(descs) else ""
        h = hrs[i] if i < len(hrs) else ""
        c = counts[i] if i < len(counts) else ""
        if p or d or h or c:
            raw.append((p, d, h, c))
    if not raw:
        raw = [("", "", "", "")]

    parsed_hrs = []
    for _, _, h, _ in raw:
        try:
            parsed_hrs.append(float(h) if h else 0.0)
        except ValueError:
            parsed_hrs.append(0.0)
    total_hr = sum(parsed_hrs)
    day_pct = round((total_hr / WORK_HOURS_PER_DAY) * 100) if WORK_HOURS_PER_DAY > 0 else 0

    rows = []
    for idx, (p, d, h, c) in enumerate(raw):
        hr_val = parsed_hrs[idx]
        try:
            count_val = float(c) if c else 0.0
        except ValueError:
            count_val = 0.0
        pct = round((hr_val / total_hr) * 100) if total_hr > 0 else 0
        target = target_lookup.get(p, {})
        target_count_hr = target.get("Target_Count_Hr") or ""
        target_hr_pct = ""
        try:
            tch = float(target_count_hr)
            if tch > 0 and hr_val > 0:
                target_hr_pct = round((count_val / (tch * hr_val)) * 100)
        except (ValueError, TypeError):
            target_hr_pct = ""

        sub = dict(record)
        sub["Process"] = p
        sub["Description"] = d
        sub["Hr"] = h
        sub["Count"] = c
        sub["_pct"] = pct
        sub["_day_pct"] = day_pct
        sub["_target_hr"] = target.get("Target_Hr") or ""
        sub["_target_pct"] = target.get("Target_Pct") or ""
        sub["_target_count_hr"] = target_count_hr
        sub["_target_hr_pct"] = target_hr_pct
        sub["_is_first"] = (idx == 0)
        sub["_group_size"] = len(raw)
        rows.append(sub)
    return rows


REPORT_COLUMNS = (
    [label for _, label in TRACKER_FIELDS[:8]]      # ... up through "Hr"
    + ["Day Overall %"]
    + [label for _, label in TRACKER_FIELDS[8:]]    # Description, Logged By, Count
    + ["% of Day", "Target hr%"]
)
REPORT_COLUMN_WIDTHS = [12, 8, 10, 16, 16, 24, 10, 8, 14, 24, 22, 10, 10, 12]
REPORT_HR_INSERT_AT = 8          # 0-based position where Day Overall % lands
REPORT_LOGGED_BY_COLUMN = "Logged By"


def build_report_sheet(wb, date_str, target_lookup, logged_by_filter=None):
    """Add one sheet named `date_str` to `wb`, formatted like the admin
    panel's Records table: one row per process (entries with more than one
    process are split), a 'Day Overall %' column right after Hr (the
    entry's total logged Hr as a % of the 8-hr work day), and '% of Day' /
    'Target hr%' columns at the end. The Logged By column is kept in the
    data but hidden, so it isn't shown unless a viewer un-hides it.

    If `logged_by_filter` is given, only that user's own entries for the
    date are included — used for the "download my data" button on the
    regular user's page, so each user can export just their own records
    in the exact same report format admin downloads use. Returns True if
    the sheet ended up with at least one data row, False otherwise (so
    callers can skip/remove empty sheets, e.g. dates with no data for
    the filtered user)."""
    ws = wb.create_sheet(title=date_str)
    ws.append(REPORT_COLUMNS)
    style_header(ws, REPORT_COLUMNS, REPORT_COLUMN_WIDTHS)

    records = get_records_for_date(date_str)
    if logged_by_filter is not None:
        records = [r for r in records if r.get("Logged_By") == logged_by_filter]

    has_rows = False
    for record in records:
        for sub in expand_record_rows(record, target_lookup):
            base_values = [sub.get(key, "") for key, _ in TRACKER_FIELDS]
            day_pct = sub.get("_day_pct")
            row = (
                base_values[:REPORT_HR_INSERT_AT]
                + [f"{day_pct}%" if day_pct != "" else ""]
                + base_values[REPORT_HR_INSERT_AT:]
            )
            row.append(f"{sub['_pct']}%" if sub.get("_pct") != "" else "")
            target_hr_pct = sub.get("_target_hr_pct")
            row.append(f"{target_hr_pct}%" if target_hr_pct != "" else "")
            ws.append(row)
            row_idx = ws.max_row
            for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).font = CELL_FONT
            has_rows = True

    logged_by_col = REPORT_COLUMNS.index(REPORT_LOGGED_BY_COLUMN) + 1
    ws.column_dimensions[get_column_letter(logged_by_col)].hidden = True
    return has_rows


def build_report_workbook(dates, logged_by_filter=None):
    """Build a fresh workbook with one report-style sheet per date in
    `dates` (see build_report_sheet), plus the Master and Process_List
    sheets kept at the end for reference. No Users sheet is ever included.

    If `logged_by_filter` is set (a user's email), each date sheet only
    contains that user's own entries, dates with none of their entries are
    dropped entirely, and the Master/Process_List reference sheets are left
    out — this is the "download my data" case for the regular user page,
    as opposed to the admin's full/monthly downloads."""
    wb = Workbook()
    wb.remove(wb.active)

    processes = get_process_list()
    target_lookup = {p["Process"]: p for p in processes if p.get("Process")}

    any_rows = False
    for date_str in sorted(dates):
        has_rows = build_report_sheet(wb, date_str, target_lookup, logged_by_filter)
        if logged_by_filter is not None and not has_rows:
            del wb[date_str]
        else:
            any_rows = any_rows or has_rows

    if not dates or (logged_by_filter is not None and not any_rows):
        ws = wb.create_sheet("Info")
        ws["A1"] = "No data available for the selected range."

    if logged_by_filter is not None:
        return wb

    # Master + Process_List sheets, pulled live from the database (kept in
    # the download for reference; Users/login accounts are never included).
    for sheet_name, columns, widths, records in (
        (MASTER_SHEET, MASTER_COLUMNS, [10, 12, 20], get_master_list()),
        (PROCESS_SHEET, PROCESS_COLUMNS, [30, 14, 14, 16], processes),
    ):
        dst = wb.create_sheet(sheet_name)
        dst.append(columns)
        style_header(dst, columns, widths)
        for record in records:
            row = [record.get(col, "") for col in columns]
            dst.append(row)
            row_idx = dst.max_row
            for col_idx in range(1, len(columns) + 1):
                dst.cell(row=row_idx, column=col_idx).font = CELL_FONT

    return wb


def entry_counts_by_user(records):
    """Per Logged_By user: number of entries submitted + total Hr, for a
    given date's records. Sorted by count, highest first."""
    counts = {}
    for r in records:
        who = r.get("Logged_By") or "(unknown)"
        if who not in counts:
            counts[who] = {"Logged_By": who, "count": 0, "hr": 0.0}
        counts[who]["count"] += 1
        counts[who]["hr"] += parse_hr_total(r.get("Hr"))
    return sorted(counts.values(), key=lambda x: x["count"], reverse=True)


# ---------------------------------------------------------------------------
# Email reminder (3:00 PM daily, for anyone under 8 hrs today)
# ---------------------------------------------------------------------------
def send_email(to_email, subject, body):
    if not SMTP_SERVER:
        print(f"[reminder] SMTP not configured — would have emailed {to_email}: {subject}")
        return
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = to_email
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            if SMTP_USERNAME and SMTP_PASSWORD:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, [to_email], msg.as_string())
    except Exception as exc:  # pragma: no cover - best-effort reminder
        print(f"[reminder] Failed to email {to_email}: {exc}")


def send_daily_pending_hr_reminders():
    """For every login account, if today's logged Hr total is below the
    8 hr/day target (or nothing was filled at all), email that account's
    login address a reminder."""
    ensure_workbook()
    today = now().strftime("%Y-%m-%d")
    today_records = get_records_for_date(today)
    for user in get_users():
        email = user.get("Email")
        if not email:
            continue
        my_records = [r for r in today_records if r.get("Logged_By") == email]
        pending = pending_hr_for_records(my_records)
        if pending > 0:
            worked = total_hr_for_records(my_records)
            body = (
                f"Hi {user.get('Name') or ''},\n\n"
                f"As of 3:00 PM, you have logged {worked:g} of the {WORK_HOURS_PER_DAY:g} "
                f"working hours required for {today} in the Productivity Tracker.\n"
                f"Pending: {pending:g} hr.\n\n"
                f"Please log in and fill in the remaining entries.\n"
            )
            send_email(email, f"Productivity Tracker — {pending:g} hr pending for {today}", body)


BACKUP_FOLDER = os.path.join(DATA_FOLDER, "backups")


def backup_all_data_to_excel():
    """Dump every table (Users, Master, Process_List, and every date's raw
    tracker entries) into a single timestamped .xlsx file under
    BACKUP_FOLDER. This exists because the live data now lives in a
    database, not an Excel file, so this is a periodic safety copy of the
    old-style 'every entry, every date' data in case the database is ever
    lost or needs auditing. Keeps at most the 30 most recent backup files
    (older ones are deleted) so the backups folder doesn't grow forever.
    Returns the path of the file that was written."""
    ensure_workbook()
    os.makedirs(BACKUP_FOLDER, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, columns, widths, records in (
        (USERS_SHEET, USER_COLUMNS, [30, 18, 20], get_users()),
        (MASTER_SHEET, MASTER_COLUMNS, [10, 12, 20], get_master_list()),
        (PROCESS_SHEET, PROCESS_COLUMNS, [30, 14, 14, 16], get_process_list()),
    ):
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        style_header(ws, columns, widths)
        for record in records:
            row = [record.get(col, "") for col in columns]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).font = CELL_FONT

    for date_str in list_available_dates():
        ws = wb.create_sheet(title=date_str)
        ws.append(TRACKER_COLUMNS)
        style_header(ws, TRACKER_COLUMNS, [12, 8, 10, 16, 16, 24, 10, 8, 24, 22, 10])
        for record in get_records_for_date(date_str):
            row = [record.get(col, "") for col in TRACKER_COLUMNS]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx in range(1, len(TRACKER_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).font = CELL_FONT

    timestamp = now().strftime("%Y-%m-%d_%H%M%S")
    path = os.path.join(BACKUP_FOLDER, f"tracker_backup_{timestamp}.xlsx")
    wb.save(path)

    # Keep only the 30 most recent backups.
    existing = sorted(
        f for f in os.listdir(BACKUP_FOLDER) if f.startswith("tracker_backup_") and f.endswith(".xlsx")
    )
    for old_file in existing[:-30]:
        try:
            os.remove(os.path.join(BACKUP_FOLDER, old_file))
        except OSError:
            pass

    return path


def list_backup_files():
    """Backup filenames under BACKUP_FOLDER, newest first."""
    if not os.path.exists(BACKUP_FOLDER):
        return []
    files = [f for f in os.listdir(BACKUP_FOLDER) if f.startswith("tracker_backup_") and f.endswith(".xlsx")]
    return sorted(files, reverse=True)


def _start_reminder_scheduler():
    """Best-effort background scheduler for the 3:00 PM reminder. Uses
    APScheduler if installed; silently no-ops otherwise so the app still
    runs without the extra dependency."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        print("[reminder] apscheduler not installed — daily 3:00 PM reminder disabled. "
              "Run `pip install apscheduler` to enable it.")
        return
    scheduler = BackgroundScheduler(daemon=True, timezone=ZoneInfo(APP_TIMEZONE))
    scheduler.add_job(
        send_daily_pending_hr_reminders,
        "cron", hour=REMINDER_HOUR, minute=REMINDER_MINUTE,
    )
    scheduler.add_job(
        backup_all_data_to_excel,
        "cron", hour=23, minute=55,
    )
    scheduler.start()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)
    return wrapper


def user_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_email"):
            return redirect(url_for("user_login"))
        return view_func(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------------------------
# HTML templates
# ---------------------------------------------------------------------------
BASE_STYLE = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Manrope:wght@500;700;800&family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
    :root {
        --ink: #1B2A3D; --ink-soft: #33445A; --line: #DCDFD8;
        --primary: #D98E2B; --primary-dark: #B8741C; --primary-tint: #FBEFDD;
        --amber: #D98E2B; --amber-tint: #FBEFDD;
        --green: #3F7268; --green-tint: #E7F0EE;
        --red: #B94A3D; --red-tint: #FBEAE7;
        --surface: #FFFFFF; --canvas: #F6F5F1;
    }
    * { box-sizing: border-box; }
    body { font-family: 'Segoe UI', 'Inter', -apple-system, BlinkMacSystemFont, Arial, sans-serif; max-width: 1000px; margin: 28px auto; background: var(--canvas); color: #20262E; }
    h1 { font-family: 'Manrope', Arial, sans-serif; font-weight: 800; color: var(--ink); letter-spacing: -0.01em; margin: 0; }
    h2 { font-family: 'Manrope', Arial, sans-serif; font-weight: 700; color: var(--ink); font-size: 16px; margin: 0 0 4px; }
    .card { background: var(--surface); padding: 22px 24px; border-radius: 4px; border: 1px solid var(--line); margin-bottom: 20px; }
    .card-section { border-left: 3px solid var(--primary); padding-left: 18px; margin-top: 18px; }
    .card-section:first-of-type { margin-top: 4px; }
    .section-title { font-family: 'Manrope', Arial, sans-serif; font-weight: 700; font-size: 12px; color: var(--green);
        text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 10px; }
    label { display: block; margin-top: 12px; font-weight: 600; font-size: 13px; color: #5C6773; }
    input[type=text], input[type=date], input[type=number], input[type=password], input[type=email], select {
        width: 100%; padding: 9px 10px; margin-top: 5px; border: 1px solid var(--line); border-radius: 3px;
        box-sizing: border-box; font-family: 'Segoe UI', 'Inter', Arial, sans-serif; font-size: 14px; color: #20262E; background: #fff;
    }
    input:focus, select:focus { outline: 2px solid var(--primary); outline-offset: 1px; border-color: var(--primary); }
    .row2 { display: grid; grid-template-columns: 1fr 1fr; gap: 0 16px; }
    .row3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 0 16px; }
    .row4 { display: grid; grid-template-columns: 1.4fr 1fr 0.7fr 0.7fr; gap: 0 16px; }
    button, .btn { margin-top: 18px; padding: 9px 20px; background: var(--primary); color: #2b1c07; border: none;
        border-radius: 3px; font-size: 14px; font-weight: 700; cursor: pointer; text-decoration: none; display: inline-block;
        font-family: 'Segoe UI', 'Inter', Arial, sans-serif; }
    button:hover, .btn:hover { background: var(--primary-dark); }
    .btn-danger { background: #fff; color: var(--red); border: 1px solid var(--red); }
    .btn-danger:hover { background: var(--red); color: #fff; }
    .btn-small { padding: 5px 12px; font-size: 12px; margin: 0 4px 0 0; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; }
    th, td { border: none; border-bottom: 1px solid var(--line); padding: 8px 10px; font-size: 13px; text-align: left; }
    th { background: var(--ink); color: #fff; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.03em; }
    tbody tr:nth-child(even) { background: #FAFAF7; }
    tbody tr:hover { background: #F1EFE7; }
    .flash { padding: 10px 14px; background: var(--green-tint); color: var(--green); border-radius: 3px; margin-bottom: 16px; font-weight: 500; }
    .flash-error { background: var(--red-tint); color: var(--red); }
    .note { font-size: 13px; color: #5C6773; }
    .topbar { display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 8px;
        background: var(--ink); color: #fff; border-radius: 4px; padding: 16px 22px; margin-bottom: 20px; }
    .topbar h1 { color: #fff; }
    .topbar a { color: #fff; text-decoration: none; font-size: 13px; margin-left: 12px; font-weight: 600; border: 1px solid #3E5065; padding: 5px 10px; border-radius: 3px; }
    .topbar a:hover { background: #2A3A4E; }
    .tag { display: inline-block; background: var(--primary); color: #2b1c07; padding: 3px 10px; border-radius: 10px; font-size: 11px; font-weight: 700; letter-spacing: .02em; }
    .proc-target { font-size: 12px; color: var(--green); font-weight: 600; margin-top: 4px; min-height: 16px; }
    body.login-page { max-width: 380px; margin: 0 auto; min-height: 100vh; display: flex; flex-direction: column;
        justify-content: center; padding: 16px; background: var(--canvas); }
    body.login-page h1 { text-align: center; font-size: 20px; margin: 0 0 14px; color: var(--ink); }
    body.login-page .card { padding: 20px 22px; margin-bottom: 14px; border-radius: 4px; box-shadow: 0 12px 40px rgba(27,42,61,0.18); }
    body.login-page p { text-align: center; margin: 0; }
    body.login-page p a { color: var(--ink); font-weight: 600; }
</style>
"""

FLASHES = """
{% with messages = get_flashed_messages(with_categories=true) %}
  {% if messages %}
    {% for cat, m in messages %}
      <div class="flash {% if cat == 'error' %}flash-error{% endif %}">{{ m }}</div>
    {% endfor %}
  {% endif %}
{% endwith %}
"""

# --- User: login -----------------------------------------------------------
USER_LOGIN_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>User Login</title>""" + BASE_STYLE + """</head>
<body class="login-page">
    <h1>👤 User Login</h1>
    """ + FLASHES + """
    <div class="card">
        <p class="note">Log in with the email &amp; password given to you by your admin.
        You may use this login to enter data for <b>any</b> employee in the list — it does not have to be your own record.</p>
        <form method="POST">
            <label>Email</label>
            <input type="email" name="email" required autofocus>
            <label>Password</label>
            <input type="password" name="password" required>
            <button type="submit">Login</button>
        </form>
    </div>
    <p><a href="{{ url_for('admin_login') }}">🔐 Admin login instead</a></p>
</body>
</html>
"""

# --- User: form + own submissions ------------------------------------------
USER_HOME_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Productivity Tracker</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>📊 LN Risk Map — Productivity Tracker</h1>
        <div>
            <span class="tag">Logged in as {{ user_name }} ({{ user_email }})</span>
            <a href="{{ url_for('user_logout') }}">Log out</a>
        </div>
    </div>
    """ + FLASHES + """

    <div class="card">
        <h2>New Entry</h2>
        <form method="POST" action="{{ url_for('submit') }}">

            <div class="card-section">
                <div class="section-title">Employee</div>
                <div class="row2">
                    <div>
                        <label>Date</label>
                        <input type="date" name="Date" value="{{ today }}" required>
                    </div>
                    <div>
                        <label>Select Employee (from master list)</label>
                        <select id="empSelect" onchange="fillEmp()">
                            <option value="">-- choose employee --</option>
                            {% for m in master %}
                            <option value="{{ loop.index0 }}"
                                data-band="{{ m.Band or '' }}"
                                data-empid="{{ m.Emp_Id or '' }}"
                                data-empname="{{ m.Emp_Name or '' }}">
                                {{ m.Emp_Id }} — {{ m.Emp_Name }}
                            </option>
                            {% endfor %}
                        </select>
                    </div>
                </div>

                <div class="row3">
                    <div><label>Band</label><input type="text" name="Band" id="Band"></div>
                    <div><label>Emp_Id</label><input type="text" name="Emp_Id" id="Emp_Id"></div>
                    <div><label>Emp_Name</label><input type="text" name="Emp_Name" id="Emp_Name"></div>
                </div>
            </div>

            <div style="background:var(--amber-tint);border:1px solid #ecd9a0;border-radius:8px;padding:12px 16px;margin-top:18px;">
                <label style="display:flex;align-items:center;gap:8px;cursor:pointer;margin-top:0;color:var(--ink);">
                    <input type="checkbox" id="leaveCheck" onchange="toggleLeave()" style="width:auto;margin-top:0;">
                    On Leave today (auto-fills {{ '%g'|format(leave_hr) }} hr)
                </label>
            </div>

            <div id="processSection" class="card-section">
                <div class="section-title">Process &amp; Hours</div>
                <div id="processRows">
                    <div class="row4 process-row">
                        <div>
                            <select class="proc-select" onchange="updateProcTarget(this)">
                                <option value="">-- select process --</option>
                                {% for p in processes %}
                                <option value="{{ p.Process }}" data-target-hr="{{ p.Target_Hr or '' }}" data-target-pct="{{ p.Target_Pct or '' }}" data-target-count-hr="{{ p.Target_Count_Hr or '' }}">{{ p.Process }}</option>
                                {% endfor %}
                            </select>
                            <div class="proc-target note"></div>
                        </div>
                        <div><input type="text" class="proc-desc" placeholder="Description"></div>
                        <div><input type="number" step="0.25" class="proc-hr" placeholder="Hr"></div>
                        <div><input type="number" step="1" class="proc-count" placeholder="Count"></div>
                    </div>
                </div>
                <button type="button" class="btn btn-small" onclick="addProcessRow()">+ Add another process row</button>
            </div>
            <input type="hidden" name="Process" id="Process_hidden">
            <input type="hidden" name="Description" id="Description_hidden">
            <input type="hidden" name="Hr" id="Hr_hidden">
            <input type="hidden" name="Count" id="Count_hidden">

            <div class="card-section">
                <div class="section-title">Additional Notes</div>
                <div class="row2">
                    <div><label>Other</label><input type="text" name="Other"></div>
                    <div><label>Description</label><input type="text" name="Other_Description"></div>
                </div>
            </div>

            <button type="submit" onclick="return collectProcessRows()">Save Entry</button>
        </form>
    </div>

    <div class="card" style="{% if pending_hr > 0 %}border:2px solid var(--red);{% else %}border:2px solid var(--green);{% endif %}">
        <h2>Today's Hours — {{ today }}</h2>
        <div class="row3">
            <div><span class="tag">Target: {{ '%g'|format(target_hr) }} hr/day</span></div>
            <div><span class="tag">Logged: {{ '%g'|format(total_hr) }} hr</span></div>
            <div><span class="tag" style="{% if pending_hr > 0 %}background:var(--red-tint);color:var(--red);{% else %}background:var(--green-tint);color:var(--green);{% endif %}">
                Pending: {{ '%g'|format(pending_hr) }} hr
            </span></div>
        </div>
        {% if pending_hr > 0 %}
        <p style="color:var(--red);font-weight:700;margin-top:14px;margin-bottom:0;">
            ⚠ Productivity not complete — {{ '%g'|format(pending_hr) }} hr still pending out of {{ '%g'|format(target_hr) }} hr/day.
        </p>
        {% else %}
        <p style="color:var(--green);font-weight:700;margin-top:14px;margin-bottom:0;">
            ✅ {{ '%g'|format(target_hr) }} hr/day completed.
        </p>
        {% endif %}
    </div>

    <div class="card">
        <h2>Your Submissions — {{ today }}</h2>
        <p class="note">You can edit or delete entries you submitted today. Entries with more than one process are split into one row per process below.</p>
        {% if records %}
        <table>
            <tr>{% for _, label in fields %}<th>{{ label }}</th>{% endfor %}<th>% of Day</th><th>Target hr%</th><th>Actions</th></tr>
            {% for r in records %}
            {% for sub in r._subrows %}
            <tr>
                {% for key, _ in fields %}
                    {% if key in ('Process', 'Description', 'Hr', 'Count') %}
                    <td>{{ sub[key] }}</td>
                    {% elif sub._is_first %}
                    <td{% if sub._group_size > 1 %} rowspan="{{ sub._group_size }}"{% endif %}>{{ r[key] }}</td>
                    {% endif %}
                {% endfor %}
                <td>{{ sub._pct }}%</td>
                <td>
                    {% if sub._target_hr or sub._target_pct %}<span class="note">target {{ sub._target_hr or '-' }}hr ({{ sub._target_pct or '-' }}%)</span>{% endif %}
                    {% if sub._target_hr_pct != '' %}<span class="note">{{ sub._target_hr_pct }}%</span>{% endif %}
                </td>
                {% if sub._is_first %}
                <td{% if sub._group_size > 1 %} rowspan="{{ sub._group_size }}"{% endif %}>
                    <a class="btn btn-small" href="{{ url_for('user_edit', date=today, row=r['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('user_delete', date=today, row=r['_row']) }}"
                          onsubmit="return confirm('Delete this entry?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
                {% endif %}
            </tr>
            {% endfor %}
            {% endfor %}
        </table>
        {% else %}
        <p>No entries submitted yet today.</p>
        {% endif %}
    </div>

<script>
function fillEmp() {
    var sel = document.getElementById('empSelect');
    var opt = sel.options[sel.selectedIndex];
    if (!opt || !opt.value) return;
    document.getElementById('Band').value = opt.getAttribute('data-band');
    document.getElementById('Emp_Id').value = opt.getAttribute('data-empid');
    document.getElementById('Emp_Name').value = opt.getAttribute('data-empname');
}

function toggleLeave() {
    var checked = document.getElementById('leaveCheck').checked;
    document.getElementById('processSection').style.display = checked ? 'none' : '';
}

function updateProcTarget(select) {
    var opt = select.options[select.selectedIndex];
    var targetDiv = select.parentElement.querySelector('.proc-target');
    var hr = opt.getAttribute('data-target-hr');
    var pct = opt.getAttribute('data-target-pct');
    var countHr = opt.getAttribute('data-target-count-hr');
    if (opt.value && (hr || pct || countHr)) {
        var parts = [];
        if (hr || pct) parts.push((hr || '-') + ' hr / ' + (pct || '-') + '%');
        if (countHr) parts.push(countHr + ' /hr target');
        targetDiv.textContent = 'Target: ' + parts.join(' · ');
    } else {
        targetDiv.textContent = '';
    }
}

function addProcessRow() {
    var container = document.getElementById('processRows');
    var row = container.children[0].cloneNode(true);
    row.querySelector('.proc-select').selectedIndex = 0;
    row.querySelector('.proc-desc').value = '';
    row.querySelector('.proc-hr').value = '';
    row.querySelector('.proc-count').value = '';
    row.querySelector('.proc-target').textContent = '';
    container.appendChild(row);
}

function collectProcessRows() {
    if (document.getElementById('leaveCheck').checked) {
        document.getElementById('Process_hidden').value = 'Leave';
        document.getElementById('Description_hidden').value = 'Leave';
        document.getElementById('Hr_hidden').value = '{{ leave_hr }}';
        document.getElementById('Count_hidden').value = '';
        return true;
    }
    var selects = document.querySelectorAll('#processRows .proc-select');
    var descs = document.querySelectorAll('#processRows .proc-desc');
    var hrs = document.querySelectorAll('#processRows .proc-hr');
    var counts = document.querySelectorAll('#processRows .proc-count');
    var procs = [], descVals = [], hrVals = [], countVals = [];
    for (var i = 0; i < selects.length; i++) {
        var p = selects[i].value.trim();
        var d = descs[i].value.trim();
        var h = hrs[i].value.trim();
        var c = counts[i].value.trim();
        if (p || d || h || c) {
            procs.push(p);
            descVals.push(d);
            hrVals.push(h);
            countVals.push(c);
        }
    }
    document.getElementById('Process_hidden').value = procs.join('; ');
    document.getElementById('Description_hidden').value = descVals.join('; ');
    document.getElementById('Hr_hidden').value = hrVals.join('; ');
    document.getElementById('Count_hidden').value = countVals.join('; ');
    if (procs.length === 0) {
        alert('Please select at least one process.');
        return false;
    }
    return true;
}
</script>
</body>
</html>
"""

# --- Admin: login ------------------------------------------------------------
ADMIN_LOGIN_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Admin Login</title>""" + BASE_STYLE + """</head>
<body class="login-page">
    <h1>🔐 Admin Login</h1>
    """ + FLASHES + """
    <div class="card">
        <form method="POST">
            <label>Username</label>
            <input type="text" name="username" required autofocus>
            <label>Password</label>
            <input type="password" name="password" required>
            <button type="submit">Login</button>
        </form>
    </div>
    <p><a href="{{ url_for('user_login') }}">👤 User login instead</a></p>
</body>
</html>
"""

# --- Admin: main panel (records) --------------------------------------------
ADMIN_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Admin Panel</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>🛠 Admin Panel</h1>
        <div>
            <a href="{{ url_for('admin_master') }}">Employee Master List</a>
            <a href="{{ url_for('admin_process') }}">Process List</a>
            <a href="{{ url_for('admin_users') }}">Login Accounts</a>
            <a href="{{ url_for('admin_logout') }}">Log out</a>
        </div>
    </div>
    """ + FLASHES + """

    <div class="card">
        <form method="GET" action="{{ url_for('admin_panel') }}">
            <label>Select date</label>
            <select name="date" onchange="this.form.submit()">
                {% for d in dates %}
                <option value="{{ d }}" {% if d == selected_date %}selected{% endif %}>{{ d }}</option>
                {% endfor %}
            </select>
        </form>
        <a class="btn" href="{{ url_for('download_excel') }}">⬇ Download Full Excel File</a>

        {% if months %}
        <form method="GET" action="{{ url_for('admin_panel') }}" style="display:inline">
            <input type="hidden" name="date" value="{{ selected_date }}">
            <label>Select month</label>
            <select name="month" onchange="this.form.submit()">
                {% for m in months %}
                <option value="{{ m }}" {% if m == selected_month %}selected{% endif %}>{{ m }}</option>
                {% endfor %}
            </select>
        </form>
        <a class="btn" href="{{ url_for('download_excel_month', month=selected_month) }}">⬇ Download Month's Excel File</a>
        {% endif %}
    </div>

    <div class="card">
        <h2>Backups</h2>
        <p class="note">A full data backup is saved automatically every night. You can also trigger one now, or download a past one.</p>
        <form method="POST" action="{{ url_for('backup_now') }}" style="display:inline">
            <button type="submit">📦 Backup Now</button>
        </form>
        {% if backups %}
        <form method="GET" action="" style="display:inline" onsubmit="return false;">
            <select id="backupSelect">
                {% for b in backups %}
                <option value="{{ b }}">{{ b }}</option>
                {% endfor %}
            </select>
        </form>
        <a class="btn btn-small" href="#" onclick="downloadSelectedBackup(); return false;">⬇ Download Selected Backup</a>
        <script>
        function downloadSelectedBackup() {
            var sel = document.getElementById('backupSelect');
            if (sel && sel.value) {
                window.location.href = "/admin/backup/download/" + encodeURIComponent(sel.value);
            }
        }
        </script>
        {% else %}
        <p>No backups yet — click "Backup Now" to create the first one.</p>
        {% endif %}
    </div>

    <div class="card">
        <h2>Add Entry (as any user)</h2>
        <p class="note">Log an entry on someone else's behalf, for today or a past date — e.g. add {{ today }}'s or an earlier date's data for a user who missed it.</p>
        <form method="POST" action="{{ url_for('admin_add_entry') }}">
            <div class="card-section">
                <div class="section-title">Employee</div>
                <div class="row2">
                    <div>
                        <label>Date</label>
                        <input type="date" name="Date" value="{{ selected_date }}" required>
                    </div>
                    <div>
                        <label>Log entry for (user)</label>
                        <select name="Logged_By" required>
                            <option value="">-- choose user --</option>
                            {% for u in users %}
                            <option value="{{ u.Email }}">{{ u.Name or u.Email }} ({{ u.Email }})</option>
                            {% endfor %}
                        </select>
                    </div>
                </div>
                <div class="row2">
                    <div>
                        <label>Select Employee (from master list)</label>
                        <select id="adminEmpSelect" onchange="adminFillEmp()">
                            <option value="">-- choose employee --</option>
                            {% for m in master %}
                            <option value="{{ loop.index0 }}"
                                data-band="{{ m.Band or '' }}"
                                data-empid="{{ m.Emp_Id or '' }}"
                                data-empname="{{ m.Emp_Name or '' }}">
                                {{ m.Emp_Id }} — {{ m.Emp_Name }}
                            </option>
                            {% endfor %}
                        </select>
                    </div>
                </div>
                <div class="row3">
                    <div><label>Band</label><input type="text" name="Band" id="adminBand"></div>
                    <div><label>Emp_Id</label><input type="text" name="Emp_Id" id="adminEmpId"></div>
                    <div><label>Emp_Name</label><input type="text" name="Emp_Name" id="adminEmpName"></div>
                </div>
            </div>

            <div id="adminProcessSection" class="card-section">
                <div class="section-title">Process &amp; Hours</div>
                <div id="adminProcessRows">
                    <div class="row4 admin-process-row">
                        <div>
                            <select class="admin-proc-select">
                                <option value="">-- select process --</option>
                                {% for p in processes %}
                                <option value="{{ p.Process }}">{{ p.Process }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div><input type="text" class="admin-proc-desc" placeholder="Description"></div>
                        <div><input type="number" step="0.25" class="admin-proc-hr" placeholder="Hr"></div>
                        <div><input type="number" step="1" class="admin-proc-count" placeholder="Count"></div>
                    </div>
                </div>
                <button type="button" class="btn btn-small" onclick="addAdminProcessRow()">+ Add another process row</button>
            </div>
            <input type="hidden" name="Process" id="adminProcess_hidden">
            <input type="hidden" name="Description" id="adminDescription_hidden">
            <input type="hidden" name="Hr" id="adminHr_hidden">
            <input type="hidden" name="Count" id="adminCount_hidden">

            <div class="card-section">
                <div class="section-title">Additional Notes</div>
                <div class="row2">
                    <div><label>Other</label><input type="text" name="Other"></div>
                    <div><label>Description</label><input type="text" name="Other_Description"></div>
                </div>
            </div>

            <button type="submit" onclick="return collectAdminProcessRows()">Save Entry</button>
        </form>
    </div>

    <script>
    function adminFillEmp() {
        var sel = document.getElementById('adminEmpSelect');
        var opt = sel.options[sel.selectedIndex];
        if (!opt || !opt.value) return;
        document.getElementById('adminBand').value = opt.getAttribute('data-band');
        document.getElementById('adminEmpId').value = opt.getAttribute('data-empid');
        document.getElementById('adminEmpName').value = opt.getAttribute('data-empname');
    }

    function addAdminProcessRow() {
        var container = document.getElementById('adminProcessRows');
        var row = container.children[0].cloneNode(true);
        row.querySelectorAll('input').forEach(function(el) { el.value = ''; });
        row.querySelectorAll('select').forEach(function(el) { el.selectedIndex = 0; });
        container.appendChild(row);
    }

    function collectAdminProcessRows() {
        var rows = document.querySelectorAll('#adminProcessRows .admin-process-row');
        var procs = [], descVals = [], hrVals = [], countVals = [];
        rows.forEach(function(row) {
            var p = row.querySelector('.admin-proc-select').value;
            var d = row.querySelector('.admin-proc-desc').value;
            var h = row.querySelector('.admin-proc-hr').value;
            var c = row.querySelector('.admin-proc-count').value;
            if (p) {
                procs.push(p);
                descVals.push(d);
                hrVals.push(h);
                countVals.push(c);
            }
        });
        document.getElementById('adminProcess_hidden').value = procs.join('; ');
        document.getElementById('adminDescription_hidden').value = descVals.join('; ');
        document.getElementById('adminHr_hidden').value = hrVals.join('; ');
        document.getElementById('adminCount_hidden').value = countVals.join('; ');
        if (procs.length === 0) {
            alert('Please select at least one process.');
            return false;
        }
        return true;
    }
    </script>

    <div class="card">
        <h2>Productivity Count — {{ selected_date }}</h2>
        <p class="note">Number of tracker entries each user submitted on this date.</p>
        {% if user_counts %}
        <table>
            <tr><th>Logged By</th><th>Entries</th><th>Total Hr</th></tr>
            {% for u in user_counts %}
            <tr>
                <td>{{ u.Logged_By }}</td>
                <td>{{ u.count }}</td>
                <td>{{ '%g'|format(u.hr) }}</td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No entries for this date.</p>
        {% endif %}
    </div>

    <div class="card">
        <h2>Records — {{ selected_date }}</h2>
        <p class="note">Entries with more than one process are split into one row per process below.</p>
        {% if records %}
        <table>
            <tr>
                {% for _, label in fields %}<th>{{ label }}</th>{% endfor %}
                <th>% of Day</th>
                <th>Target hr%</th>
                <th>Actions</th>
            </tr>
            {% for r in records %}
            {% for sub in r._subrows %}
            <tr>
                {% for key, _ in fields %}
                    {% if key in ('Process', 'Description', 'Hr', 'Count') %}
                    <td>{{ sub[key] }}</td>
                    {% elif sub._is_first %}
                    <td{% if sub._group_size > 1 %} rowspan="{{ sub._group_size }}"{% endif %}>{{ r[key] }}</td>
                    {% endif %}
                {% endfor %}
                <td>{{ sub._pct }}%</td>
                <td>
                    {% if sub._target_hr or sub._target_pct %}<span class="note">target {{ sub._target_hr or '-' }}hr ({{ sub._target_pct or '-' }}%)</span>{% endif %}
                    {% if sub._target_hr_pct != '' %}<span class="note">{{ sub._target_hr_pct }}%</span>{% endif %}
                </td>
                {% if sub._is_first %}
                <td{% if sub._group_size > 1 %} rowspan="{{ sub._group_size }}"{% endif %}>
                    <a class="btn btn-small" href="{{ url_for('admin_edit', date=selected_date, row=r['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('admin_delete', date=selected_date, row=r['_row']) }}"
                          onsubmit="return confirm('Delete this record?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
                {% endif %}
            </tr>
            {% endfor %}
            {% endfor %}
        </table>
        {% else %}
        <p>No records for this date.</p>
        {% endif %}
    </div>
</body>
</html>
"""

EDIT_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Edit Record</title>""" + BASE_STYLE + """</head>
<body>
    <h1>✏️ Edit Record — {{ date }}</h1>
    <div class="card">
        <form method="POST" onsubmit="return collectProcessRows()">
            {% for key, label in other_fields %}
            {% if key != 'Logged_By' %}
            <label>{{ label }}</label>
            <input type="text" name="{{ key }}" value="{{ record[key] or '' }}">
            {% endif %}
            {% endfor %}

            <label>Process</label>
            <div id="processRows">
                {% for row in process_rows %}
                <div class="row4 process-row">
                    <div>
                        <select class="proc-select" onchange="updateProcTarget(this)">
                            <option value="">-- select process --</option>
                            {% for p in processes %}
                            <option value="{{ p.Process }}" data-target-hr="{{ p.Target_Hr or '' }}" data-target-pct="{{ p.Target_Pct or '' }}" data-target-count-hr="{{ p.Target_Count_Hr or '' }}" {% if p.Process == row.Process %}selected{% endif %}>{{ p.Process }}</option>
                            {% endfor %}
                        </select>
                        <div class="proc-target note"></div>
                    </div>
                    <div><input type="text" class="proc-desc" placeholder="Description" value="{{ row.Description }}"></div>
                    <div><input type="number" step="0.25" class="proc-hr" placeholder="Hr" value="{{ row.Hr }}"></div>
                    <div><input type="number" step="1" class="proc-count" placeholder="Count" value="{{ row.Count }}"></div>
                </div>
                {% endfor %}
            </div>
            <button type="button" class="btn btn-small" onclick="addProcessRow()">+ Add another process row</button>
            <input type="hidden" name="Process" id="Process_hidden">
            <input type="hidden" name="Description" id="Description_hidden">
            <input type="hidden" name="Hr" id="Hr_hidden">
            <input type="hidden" name="Count" id="Count_hidden">

            <label>Logged By</label>
            {% if logged_by_readonly %}
            <input type="text" value="{{ record.Logged_By or '' }}" readonly style="background:var(--canvas);">
            <input type="hidden" name="Logged_By" value="{{ record.Logged_By or '' }}">
            {% else %}
            <input type="text" name="Logged_By" value="{{ record.Logged_By or '' }}">
            {% endif %}

            <button type="submit">Save Changes</button>
        </form>
    </div>
    <p><a href="{{ back_url }}">← {{ back_label }}</a></p>

<script>
function updateProcTarget(select) {
    var opt = select.options[select.selectedIndex];
    var targetDiv = select.parentElement.querySelector('.proc-target');
    var hr = opt.getAttribute('data-target-hr');
    var pct = opt.getAttribute('data-target-pct');
    var countHr = opt.getAttribute('data-target-count-hr');
    if (opt.value && (hr || pct || countHr)) {
        var parts = [];
        if (hr || pct) parts.push((hr || '-') + ' hr / ' + (pct || '-') + '%');
        if (countHr) parts.push(countHr + ' /hr target');
        targetDiv.textContent = 'Target: ' + parts.join(' · ');
    } else {
        targetDiv.textContent = '';
    }
}

function addProcessRow() {
    var container = document.getElementById('processRows');
    var row = container.children[0].cloneNode(true);
    row.querySelector('.proc-select').selectedIndex = 0;
    row.querySelector('.proc-desc').value = '';
    row.querySelector('.proc-hr').value = '';
    row.querySelector('.proc-count').value = '';
    row.querySelector('.proc-target').textContent = '';
    container.appendChild(row);
}

document.querySelectorAll('#processRows .proc-select').forEach(updateProcTarget);

function collectProcessRows() {
    var selects = document.querySelectorAll('#processRows .proc-select');
    var descs = document.querySelectorAll('#processRows .proc-desc');
    var hrs = document.querySelectorAll('#processRows .proc-hr');
    var counts = document.querySelectorAll('#processRows .proc-count');
    var procs = [], descVals = [], hrVals = [], countVals = [];
    for (var i = 0; i < selects.length; i++) {
        var p = selects[i].value.trim();
        var d = descs[i].value.trim();
        var h = hrs[i].value.trim();
        var c = counts[i].value.trim();
        if (p || d || h || c) {
            procs.push(p);
            descVals.push(d);
            hrVals.push(h);
            countVals.push(c);
        }
    }
    document.getElementById('Process_hidden').value = procs.join('; ');
    document.getElementById('Description_hidden').value = descVals.join('; ');
    document.getElementById('Hr_hidden').value = hrVals.join('; ');
    document.getElementById('Count_hidden').value = countVals.join('; ');
    if (procs.length === 0) {
        alert('Please select at least one process.');
        return false;
    }
    return true;
}
</script>
</body>
</html>
"""

# --- Admin: employee master list --------------------------------------------
MASTER_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Employee Master List</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>📋 Employee Master List</h1>
        <div><a href="{{ url_for('admin_panel') }}">← Admin panel</a></div>
    </div>
    """ + FLASHES + """

    <div class="card">
        <h2>Add Employee</h2>
        <form method="POST" action="{{ url_for('admin_master_add') }}">
            <div class="row3">
                <div><label>Band</label><input type="text" name="Band" required></div>
                <div><label>Emp_Id</label><input type="text" name="Emp_Id" required></div>
                <div><label>Emp_Name</label><input type="text" name="Emp_Name" required></div>
            </div>
            <button type="submit">Add to List</button>
        </form>
    </div>

    <div class="card">
        <h2>Current List</h2>
        {% if master %}
        <table>
            <tr>{% for c in columns %}<th>{{ c }}</th>{% endfor %}<th>Actions</th></tr>
            {% for m in master %}
            <tr>
                {% for c in columns %}<td>{{ m[c] }}</td>{% endfor %}
                <td>
                    <a class="btn btn-small" href="{{ url_for('admin_master_edit', row=m['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('admin_master_delete', row=m['_row']) }}"
                          onsubmit="return confirm('Remove this employee from the list?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No employees added yet.</p>
        {% endif %}
    </div>
</body>
</html>
"""

MASTER_EDIT_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Edit Employee</title>""" + BASE_STYLE + """</head>
<body>
    <h1>✏️ Edit Employee</h1>
    <div class="card">
        <form method="POST">
            {% for c in columns %}
            <label>{{ c }}</label>
            <input type="text" name="{{ c }}" value="{{ record[c] or '' }}">
            {% endfor %}
            <button type="submit">Save Changes</button>
        </form>
    </div>
    <p><a href="{{ url_for('admin_master') }}">← Back to master list</a></p>
</body>
</html>
"""

# --- Admin: process list -----------------------------------------------------
PROCESS_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Process List</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>🧩 Process List</h1>
        <div><a href="{{ url_for('admin_panel') }}">← Admin panel</a></div>
    </div>
    """ + FLASHES + """

    <div class="card">
        <h2>Add Process</h2>
        <form method="POST" action="{{ url_for('admin_process_add') }}">
            <div class="row4">
                <div><label>Process name</label><input type="text" name="Process" required></div>
                <div><label>Target Hr</label><input type="number" step="0.25" name="Target_Hr"></div>
                <div><label>Target %</label><input type="number" step="1" name="Target_Pct"></div>
                <div><label>Target Count/Hr</label><input type="number" step="1" name="Target_Count_Hr"></div>
            </div>
            <button type="submit">Add Process</button>
        </form>
        <p class="note">This is the dropdown list users choose from when filling out an entry. Target Hr / Target % are used to compare against actual logged hours in the admin records view. Target Count/Hr is the units-per-hour target used to compute "Target hr %" (Count ÷ (Target Count/Hr × Hr) × 100) for each process row.</p>
    </div>

    <div class="card">
        <h2>Current Processes</h2>
        {% if processes %}
        <table>
            <tr><th>Process</th><th>Target Hr</th><th>Target %</th><th>Target Count/Hr</th><th>Actions</th></tr>
            {% for p in processes %}
            <tr>
                <td>{{ p.Process }}</td>
                <td>{{ p.Target_Hr or '-' }}</td>
                <td>{{ p.Target_Pct or '-' }}</td>
                <td>{{ p.Target_Count_Hr or '-' }}</td>
                <td>
                    <a class="btn btn-small" href="{{ url_for('admin_process_edit', row=p['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('admin_process_delete', row=p['_row']) }}"
                          onsubmit="return confirm('Remove this process from the list?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No processes added yet.</p>
        {% endif %}
    </div>
</body>
</html>
"""

PROCESS_EDIT_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Edit Process</title>""" + BASE_STYLE + """</head>
<body>
    <h1>✏️ Edit Process</h1>
    <div class="card">
        <form method="POST">
            <label>Process name</label>
            <input type="text" name="Process" value="{{ record.Process or '' }}" required>
            <div class="row3">
                <div><label>Target Hr</label><input type="number" step="0.25" name="Target_Hr" value="{{ record.Target_Hr or '' }}"></div>
                <div><label>Target %</label><input type="number" step="1" name="Target_Pct" value="{{ record.Target_Pct or '' }}"></div>
                <div><label>Target Count/Hr</label><input type="number" step="1" name="Target_Count_Hr" value="{{ record.Target_Count_Hr or '' }}"></div>
            </div>
            <button type="submit">Save Changes</button>
        </form>
    </div>
    <p><a href="{{ url_for('admin_process') }}">← Back to process list</a></p>
</body>
</html>
"""

# --- Admin: login accounts ---------------------------------------------------
USERS_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Login Accounts</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>👥 User Login Accounts</h1>
        <div><a href="{{ url_for('admin_panel') }}">← Admin panel</a></div>
    </div>
    """ + FLASHES + """

    <div class="card">
        <h2>Add Login Account</h2>
        <form method="POST" action="{{ url_for('admin_users_add') }}">
            <div class="row3">
                <div><label>Email</label><input type="email" name="Email" required></div>
                <div><label>Password</label><input type="text" name="Password" required></div>
                <div><label>Name</label><input type="text" name="Name" required></div>
            </div>
            <p class="note">"Name" identifies the person this login belongs to. It does not
            restrict which employee's data they can fill in — that's chosen from the master list.</p>
            <button type="submit">Add Account</button>
        </form>
    </div>

    <div class="card">
        <h2>Current Accounts</h2>
        {% if users %}
        <table>
            <tr><th>Email</th><th>Password</th><th>Name</th><th>Actions</th></tr>
            {% for u in users %}
            <tr>
                <td>{{ u.Email }}</td><td>{{ u.Password }}</td><td>{{ u.Name }}</td>
                <td>
                    <a class="btn btn-small" href="{{ url_for('admin_users_edit', row=u['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('admin_users_delete', row=u['_row']) }}"
                          onsubmit="return confirm('Remove this login account?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No user accounts yet.</p>
        {% endif %}
    </div>
</body>
</html>
"""

USER_EDIT_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Edit Account</title>""" + BASE_STYLE + """</head>
<body>
    <h1>✏️ Edit Login Account</h1>
    <div class="card">
        <form method="POST">
            <label>Email</label>
            <input type="email" name="Email" value="{{ record.Email or '' }}" required>
            <label>Password</label>
            <input type="text" name="Password" value="{{ record.Password or '' }}" required>
            <label>Name</label>
            <input type="text" name="Name" value="{{ record.Name or '' }}" required>
            <button type="submit">Save Changes</button>
        </form>
    </div>
    <p><a href="{{ url_for('admin_users') }}">← Back to accounts</a></p>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# User-facing routes
# ---------------------------------------------------------------------------
@app.route("/")
def home():
    if session.get("user_email"):
        return redirect(url_for("user_home"))
    return redirect(url_for("user_login"))


@app.route("/login", methods=["GET", "POST"])
def user_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        user = find_user_by_email(email)
        if user and str(user.get("Password")) == password:
            session["user_email"] = user["Email"]
            session["user_name"] = user.get("Name") or user["Email"]
            flash("Logged in successfully.")
            return redirect(url_for("user_home"))
        flash("Invalid email or password.", "error")
    return render_template_string(USER_LOGIN_PAGE)


@app.route("/logout")
def user_logout():
    session.pop("user_email", None)
    session.pop("user_name", None)
    return redirect(url_for("user_login"))


@app.route("/tracker")
@user_required
def user_home():
    ensure_workbook()
    master = get_master_list()
    processes = get_process_list()
    records = [r for r in get_today_records() if r.get("Logged_By") == session["user_email"]]
    target_lookup = {p["Process"]: p for p in processes if p.get("Process")}
    for r in records:
        r["_subrows"] = expand_record_rows(r, target_lookup)
    today = now().strftime("%Y-%m-%d")
    total_hr = total_hr_for_records(records)
    pending_hr = pending_hr_for_records(records)
    return render_template_string(
        USER_HOME_PAGE, master=master, processes=processes, records=records, fields=TRACKER_FIELDS,
        today=today, user_name=session.get("user_name"), user_email=session.get("user_email"),
        total_hr=total_hr, pending_hr=pending_hr, target_hr=WORK_HOURS_PER_DAY, leave_hr=LEAVE_HR
    )


@app.route("/submit", methods=["POST"])
@user_required
def submit():
    data = {key: request.form.get(key, "").strip() for key, _ in TRACKER_FIELDS if key != "Logged_By"}
    data["Logged_By"] = session["user_email"]
    if not data.get("Date"):
        data["Date"] = now().strftime("%Y-%m-%d")
    save_entry(data)
    flash(f"Saved entry for {data.get('Emp_Name') or 'employee'} ({data.get('Emp_Id')}).")
    return redirect(url_for("user_home"))


@app.route("/my/edit/<date>/<int:row>", methods=["GET", "POST"])
@user_required
def user_edit(date, row):
    """Same edit flow as admin_edit, but a user may only edit their own
    submissions (Logged_By must match the logged-in user's email)."""
    records = get_records_for_date(date)
    record = next((r for r in records if r["_row"] == row), None)
    if record is None:
        abort(404)
    if record.get("Logged_By") != session["user_email"]:
        abort(403)

    if request.method == "POST":
        data = {key: request.form.get(key, "").strip() for key, _ in TRACKER_FIELDS}
        data["Logged_By"] = session["user_email"]  # can't be reassigned to someone else
        if update_record(date, row, data):
            flash("Entry updated.")
        else:
            flash("Could not update entry — sheet not found.", "error")
        return redirect(url_for("user_home"))

    processes = get_process_list()
    process_rows = split_process_rows(record)
    other_fields = [(k, l) for k, l in TRACKER_FIELDS if k not in ("Process", "Description", "Hr", "Count")]
    return render_template_string(
        EDIT_PAGE, date=date, record=record, fields=TRACKER_FIELDS,
        other_fields=other_fields, processes=processes, process_rows=process_rows,
        back_url=url_for("user_home"), back_label="Back to your entries",
        logged_by_readonly=True
    )
@app.route("/my/delete/<date>/<int:row>", methods=["POST"])
@user_required
def user_delete(date, row):
    """Same delete flow as admin_delete, but restricted to the user's own
    submissions (Logged_By must match the logged-in user's email)."""
    records = get_records_for_date(date)
    record = next((r for r in records if r["_row"] == row), None)
    if record is None:
        abort(404)
    if record.get("Logged_By") != session["user_email"]:
        abort(403)
    if delete_record(date, row):
        flash("Entry deleted.")
    else:
        flash("Could not delete entry — sheet not found.", "error")
    return redirect(url_for("user_home"))




# ---------------------------------------------------------------------------
# Admin: auth
# ---------------------------------------------------------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Logged in successfully.")
            return redirect(url_for("admin_panel"))
        flash("Invalid username or password.", "error")
    return render_template_string(ADMIN_LOGIN_PAGE)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


# ---------------------------------------------------------------------------
# Admin: tracker records
# ---------------------------------------------------------------------------
@app.route("/admin")
@admin_required
def admin_panel():
    ensure_workbook()
    dates = list_available_dates()
    today = now().strftime("%Y-%m-%d")
    selected_date = request.args.get("date") or (dates[0] if dates else today)
    records = get_records_for_date(selected_date)
    target_lookup = {p["Process"]: p for p in get_process_list() if p.get("Process")}
    for r in records:
        r["_subrows"] = expand_record_rows(r, target_lookup)
    user_counts = entry_counts_by_user(records)
    months = list_available_months()
    selected_month = request.args.get("month") or (months[0] if months else "")
    return render_template_string(
        ADMIN_PAGE, dates=dates, selected_date=selected_date,
        records=records, fields=TRACKER_FIELDS, user_counts=user_counts,
        months=months, selected_month=selected_month,
        master=get_master_list(), processes=get_process_list(), users=get_users(),
        today=today, leave_hr=LEAVE_HR, backups=list_backup_files()
    )


@app.route("/admin/add-entry", methods=["POST"])
@admin_required
def admin_add_entry():
    """Let admin add a New Entry on behalf of any user (chosen from the
    'Log entry for' dropdown) and for any date (not just today) — e.g. to
    backfill 03/09/26 while today is 04/09/26."""
    entry_date = request.form.get("Date", "").strip() or now().strftime("%Y-%m-%d")
    logged_by = request.form.get("Logged_By", "").strip()
    if not logged_by:
        flash("Please choose which user to log this entry for.", "error")
        return redirect(url_for("admin_panel", date=entry_date))

    data = {key: request.form.get(key, "").strip() for key, _ in TRACKER_FIELDS if key != "Logged_By"}
    data["Date"] = entry_date
    data["Logged_By"] = logged_by
    save_entry(data, sheet_date=entry_date)
    flash(f"Entry added for {data.get('Emp_Name') or 'employee'} on {entry_date} (logged as {logged_by}).")
    return redirect(url_for("admin_panel", date=entry_date))


@app.route("/admin/edit/<date>/<int:row>", methods=["GET", "POST"])
@admin_required
def admin_edit(date, row):
    if request.method == "POST":
        data = {key: request.form.get(key, "").strip() for key, _ in TRACKER_FIELDS}
        if update_record(date, row, data):
            flash("Record updated.")
        else:
            flash("Could not update record — sheet not found.", "error")
        return redirect(url_for("admin_panel", date=date))

    records = get_records_for_date(date)
    record = next((r for r in records if r["_row"] == row), None)
    if record is None:
        abort(404)
    processes = get_process_list()
    process_rows = split_process_rows(record)
    other_fields = [(k, l) for k, l in TRACKER_FIELDS if k not in ("Process", "Description", "Hr", "Count")]
    return render_template_string(
        EDIT_PAGE, date=date, record=record, fields=TRACKER_FIELDS,
        other_fields=other_fields, processes=processes, process_rows=process_rows,
        back_url=url_for("admin_panel", date=date), back_label="Back to admin panel",
        logged_by_readonly=False
    )


@app.route("/admin/delete/<date>/<int:row>", methods=["POST"])
@admin_required
def admin_delete(date, row):
    if delete_record(date, row):
        flash("Record deleted.")
    else:
        flash("Could not delete record — sheet not found.", "error")
    return redirect(url_for("admin_panel", date=date))


@app.route("/admin/download")
@admin_required
def download_excel():
    """Download every date's records in the same report format shown on the
    admin panel (one row per process, with % of Day / Target hr% columns).
    The Users sheet (login emails/passwords) is never included."""
    ensure_workbook()
    wb = build_report_workbook(list_available_dates())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="productivity_tracker.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/download/month/<month>")
@admin_required
def download_excel_month(month):
    """Download one month's records (e.g. month='2026-09') in the same
    report format shown on the admin panel (one row per process, with
    % of Day / Target hr% columns). Master/Process_List are kept for
    reference; Users (login emails/passwords) is never included."""
    ensure_workbook()
    month_dates = [d for d in list_available_dates() if d.startswith(month)]
    wb = build_report_workbook(month_dates)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"productivity_tracker_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/backup-now", methods=["POST"])
@admin_required
def backup_now():
    """Manually trigger the same backup the nightly job runs, so admin
    doesn't have to wait until 11:55 PM to get a fresh backup file."""
    path = backup_all_data_to_excel()
    flash(f"Backup saved: {os.path.basename(path)}")
    return redirect(url_for("admin_panel"))


@app.route("/admin/backup/download/<filename>")
@admin_required
def download_backup(filename):
    """Download a specific backup file by name. Filenames are validated
    against the actual list of backups on disk, so this can't be used to
    read arbitrary files on the server."""
    if filename not in list_backup_files():
        abort(404)
    return send_file(
        os.path.join(BACKUP_FOLDER, filename), as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------------------------
# Admin: employee master list
# ---------------------------------------------------------------------------
@app.route("/admin/master")
@admin_required
def admin_master():
    ensure_workbook()
    return render_template_string(MASTER_PAGE, master=get_master_list(), columns=MASTER_COLUMNS)


@app.route("/admin/master/add", methods=["POST"])
@admin_required
def admin_master_add():
    data = {c: request.form.get(c, "").strip() for c in MASTER_COLUMNS}
    add_master(data)
    flash(f"Added {data.get('Emp_Name')} to the master list.")
    return redirect(url_for("admin_master"))


@app.route("/admin/master/edit/<int:row>", methods=["GET", "POST"])
@admin_required
def admin_master_edit(row):
    if request.method == "POST":
        data = {c: request.form.get(c, "").strip() for c in MASTER_COLUMNS}
        update_master(row, data)
        flash("Employee updated.")
        return redirect(url_for("admin_master"))
    record = next((m for m in get_master_list() if m["_row"] == row), None)
    if record is None:
        abort(404)
    return render_template_string(MASTER_EDIT_PAGE, record=record, columns=MASTER_COLUMNS)


@app.route("/admin/master/delete/<int:row>", methods=["POST"])
@admin_required
def admin_master_delete(row):
    delete_master(row)
    flash("Employee removed from master list.")
    return redirect(url_for("admin_master"))


# ---------------------------------------------------------------------------
# Admin: process list
# ---------------------------------------------------------------------------
@app.route("/admin/process")
@admin_required
def admin_process():
    ensure_workbook()
    return render_template_string(PROCESS_PAGE, processes=get_process_list())


@app.route("/admin/process/add", methods=["POST"])
@admin_required
def admin_process_add():
    name = request.form.get("Process", "").strip()
    target_hr = request.form.get("Target_Hr", "").strip()
    target_pct = request.form.get("Target_Pct", "").strip()
    target_count_hr = request.form.get("Target_Count_Hr", "").strip()
    if name:
        add_process(name, target_hr, target_pct, target_count_hr)
        flash(f"Added process '{name}'.")
    return redirect(url_for("admin_process"))


@app.route("/admin/process/edit/<int:row>", methods=["GET", "POST"])
@admin_required
def admin_process_edit(row):
    if request.method == "POST":
        name = request.form.get("Process", "").strip()
        target_hr = request.form.get("Target_Hr", "").strip()
        target_pct = request.form.get("Target_Pct", "").strip()
        target_count_hr = request.form.get("Target_Count_Hr", "").strip()
        update_process(row, name, target_hr, target_pct, target_count_hr)
        flash("Process updated.")
        return redirect(url_for("admin_process"))
    record = next((p for p in get_process_list() if p["_row"] == row), None)
    if record is None:
        abort(404)
    return render_template_string(PROCESS_EDIT_PAGE, record=record)


@app.route("/admin/process/delete/<int:row>", methods=["POST"])
@admin_required
def admin_process_delete(row):
    delete_process(row)
    flash("Process removed.")
    return redirect(url_for("admin_process"))


# ---------------------------------------------------------------------------
# Admin: user login accounts
# ---------------------------------------------------------------------------
@app.route("/admin/users")
@admin_required
def admin_users():
    ensure_workbook()
    return render_template_string(USERS_PAGE, users=get_users())


@app.route("/admin/users/add", methods=["POST"])
@admin_required
def admin_users_add():
    email = request.form.get("Email", "").strip()
    password = request.form.get("Password", "").strip()
    name = request.form.get("Name", "").strip()
    if find_user_by_email(email):
        flash("An account with that email already exists.", "error")
    else:
        add_user(email, password, name)
        flash(f"Created login for {name} ({email}).")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/edit/<int:row>", methods=["GET", "POST"])
@admin_required
def admin_users_edit(row):
    if request.method == "POST":
        email = request.form.get("Email", "").strip()
        password = request.form.get("Password", "").strip()
        name = request.form.get("Name", "").strip()
        update_user(row, email, password, name)
        flash("Account updated.")
        return redirect(url_for("admin_users"))
    record = next((u for u in get_users() if u["_row"] == row), None)
    if record is None:
        abort(404)
    return render_template_string(USER_EDIT_PAGE, record=record)


@app.route("/admin/users/delete/<int:row>", methods=["POST"])
@admin_required
def admin_users_delete(row):
    delete_user(row)
    flash("Account removed.")
    return redirect(url_for("admin_users"))


# ---------------------------------------------------------------------------
ensure_workbook()
_start_reminder_scheduler()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
